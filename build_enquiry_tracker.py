from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT_FILE = "Hospera-Enquiry-Tracker.xlsx"


def style_header(cell):
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1B2B4B")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="D9E2F2"),
        right=Side(style="thin", color="D9E2F2"),
        top=Side(style="thin", color="D9E2F2"),
        bottom=Side(style="thin", color="D9E2F2"),
    )


def style_body(cell):
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="E6EAF2"),
        right=Side(style="thin", color="E6EAF2"),
        top=Side(style="thin", color="E6EAF2"),
        bottom=Side(style="thin", color="E6EAF2"),
    )


def main():
    workbook = Workbook()
    tracker = workbook.active
    tracker.title = "Enquiry Tracker"

    apply_sheet = workbook.create_sheet("Apply Now")
    summary = workbook.create_sheet("Dashboard")
    lists = workbook.create_sheet("Lists")

    status_options = [
        "New",
        "Attempting Contact",
        "Contacted",
        "Follow-up Due",
        "Interested",
        "Not Interested",
        "Enrolled",
        "Closed",
    ]
    priority_options = ["High", "Medium", "Low"]
    channel_options = ["Website Form", "Live Chat", "Phone", "WhatsApp", "Walk-in", "Referral", "Other"]
    outcome_options = [
        "Pending",
        "No Response",
        "Call Back Requested",
        "Meeting Scheduled",
        "Application Started",
        "Enrolled",
        "Closed Lost",
    ]
    yes_no_options = ["Yes", "No"]

    list_columns = {
        "A": ("Status", status_options),
        "B": ("Priority", priority_options),
        "C": ("Source Channel", channel_options),
        "D": ("Outcome", outcome_options),
        "E": ("Yes / No", yes_no_options),
    }

    for column, (title, values) in list_columns.items():
        lists[f"{column}1"] = title
        style_header(lists[f"{column}1"])
        for index, value in enumerate(values, start=2):
            lists[f"{column}{index}"] = value
            style_body(lists[f"{column}{index}"])
        lists.column_dimensions[column].width = 22

    lists.sheet_state = "hidden"

    enquiry_headers = [
        "Lead ID",
        "Date Received",
        "Time Received",
        "Full Name",
        "Phone",
        "Email",
        "Course / Interest",
        "Source Channel",
        "Source Page",
        "City / Country",
        "Message",
        "Status",
        "Priority",
        "Assigned To",
        "Contacted?",
        "First Contact Date",
        "Outcome",
        "Next Action",
        "Follow-up Date",
        "Last Updated",
        "Notes",
    ]

    enquiry_widths = [16, 15, 14, 24, 18, 30, 22, 18, 26, 18, 38, 18, 12, 18, 12, 17, 20, 24, 16, 16, 36]

    for column_index, (header, width) in enumerate(zip(enquiry_headers, enquiry_widths), start=1):
        cell = tracker.cell(row=1, column=column_index, value=header)
        style_header(cell)
        tracker.column_dimensions[get_column_letter(column_index)].width = width

    tracker.freeze_panes = "A2"
    tracker.auto_filter.ref = f"A1:{get_column_letter(len(enquiry_headers))}201"
    tracker.sheet_view.showGridLines = False

    for row in range(2, 202):
        tracker[f"A{row}"] = f'=IF(D{row}="","","HOS-"&TEXT(ROW()-1,"0000"))'
        tracker[f"B{row}"] = f'=IF(D{row}="","",TODAY())'
        tracker[f"C{row}"] = f'=IF(D{row}="","",TEXT(NOW(),"hh:mm"))'
        tracker[f"T{row}"] = f'=IF(D{row}="","",TODAY())'
        for column in range(1, len(enquiry_headers) + 1):
            style_body(tracker.cell(row=row, column=column))

    date_columns = ["B", "P", "S", "T"]
    for column in date_columns:
        for row in range(2, 202):
            tracker[f"{column}{row}"].number_format = "dd-mmm-yyyy"

    for row in range(2, 202):
        tracker[f"C{row}"].number_format = "hh:mm"

    validations = [
        ("H2:H201", "=Lists!$C$2:$C$8"),
        ("L2:L201", "=Lists!$A$2:$A$9"),
        ("M2:M201", "=Lists!$B$2:$B$4"),
        ("O2:O201", "=Lists!$E$2:$E$3"),
        ("Q2:Q201", "=Lists!$D$2:$D$8"),
    ]

    for cell_range, formula in validations:
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.prompt = "Choose from the drop-down list."
        validation.error = "Please select a valid option from the list."
        tracker.add_data_validation(validation)
        validation.add(cell_range)

    tracker.conditional_formatting.add(
        "L2:L201",
        FormulaRule(formula=['L2="New"'], fill=PatternFill("solid", fgColor="DDEBF7")),
    )
    tracker.conditional_formatting.add(
        "L2:L201",
        FormulaRule(formula=['L2="Follow-up Due"'], fill=PatternFill("solid", fgColor="FFF2CC")),
    )
    tracker.conditional_formatting.add(
        "L2:L201",
        FormulaRule(formula=['L2="Enrolled"'], fill=PatternFill("solid", fgColor="E2F0D9")),
    )
    tracker.conditional_formatting.add(
        "L2:L201",
        FormulaRule(formula=['L2="Closed"'], fill=PatternFill("solid", fgColor="F4CCCC")),
    )

    tracker["A204"] = "How to use"
    tracker["A204"].font = Font(bold=True, color="1B2B4B")
    tracker["A205"] = "Add every new form or chat enquiry in a new row."
    tracker["A206"] = "Use Status, Priority, Contacted?, and Outcome drop-down lists."
    tracker["A207"] = "Set Follow-up Date for anyone who needs another call or message."
    tracker["A208"] = "Use Notes for counsellor remarks, parent conversation, or eligibility details."

    apply_headers = [
        "Candidate ID",
        "Date Received",
        "Time Received",
        "Full Name",
        "Phone",
        "Email",
        "Current City",
        "Highest Qualification",
        "Interested Program",
        "Career Goal",
        "Source Channel",
        "Source Page",
        "Status",
        "Priority",
        "Assigned To",
        "Contacted?",
        "First Contact Date",
        "Outcome",
        "Next Action",
        "Follow-up Date",
        "Last Updated",
        "Notes",
    ]

    apply_widths = [18, 15, 14, 24, 18, 30, 18, 22, 28, 38, 18, 24, 18, 12, 18, 12, 17, 20, 24, 16, 16, 36]

    for column_index, (header, width) in enumerate(zip(apply_headers, apply_widths), start=1):
        cell = apply_sheet.cell(row=1, column=column_index, value=header)
        style_header(cell)
        apply_sheet.column_dimensions[get_column_letter(column_index)].width = width

    apply_sheet.freeze_panes = "A2"
    apply_sheet.auto_filter.ref = f"A1:{get_column_letter(len(apply_headers))}201"
    apply_sheet.sheet_view.showGridLines = False

    for row in range(2, 202):
        apply_sheet[f"A{row}"] = f'=IF(D{row}="","","APP-"&TEXT(ROW()-1,"0000"))'
        apply_sheet[f"B{row}"] = f'=IF(D{row}="","",TODAY())'
        apply_sheet[f"C{row}"] = f'=IF(D{row}="","",TEXT(NOW(),"hh:mm"))'
        apply_sheet[f"U{row}"] = f'=IF(D{row}="","",TODAY())'
        apply_sheet[f"N{row}"] = "High"
        apply_sheet[f"P{row}"] = "No"
        apply_sheet[f"R{row}"] = "Pending"
        for column in range(1, len(apply_headers) + 1):
            style_body(apply_sheet.cell(row=row, column=column))

    for column in ["B", "Q", "T", "U"]:
        for row in range(2, 202):
            apply_sheet[f"{column}{row}"].number_format = "dd-mmm-yyyy"

    for row in range(2, 202):
        apply_sheet[f"C{row}"].number_format = "hh:mm"

    apply_validations = [
        ("K2:K201", "=Lists!$C$2:$C$8"),
        ("M2:M201", "=Lists!$A$2:$A$9"),
        ("N2:N201", "=Lists!$B$2:$B$4"),
        ("P2:P201", "=Lists!$E$2:$E$3"),
        ("R2:R201", "=Lists!$D$2:$D$8"),
    ]

    for cell_range, formula in apply_validations:
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.prompt = "Choose from the drop-down list."
        validation.error = "Please select a valid option from the list."
        apply_sheet.add_data_validation(validation)
        validation.add(cell_range)

    apply_sheet.conditional_formatting.add(
        "N2:N201",
        FormulaRule(formula=['N2="High"'], fill=PatternFill("solid", fgColor="F4CCCC")),
    )
    apply_sheet.conditional_formatting.add(
        "M2:M201",
        FormulaRule(formula=['M2="Enrolled"'], fill=PatternFill("solid", fgColor="E2F0D9")),
    )
    apply_sheet["A204"] = "How to use"
    apply_sheet["A204"].font = Font(bold=True, color="1B2B4B")
    apply_sheet["A205"] = "Use this tab for candidates who filled Apply Now and are closer to enrollment."
    apply_sheet["A206"] = "Priority is set to High by default for faster follow-up."
    apply_sheet["A207"] = "Track counsellor ownership, first contact, and next follow-up clearly."

    summary.sheet_view.showGridLines = False
    summary["A1"] = "Hospera Enquiry Dashboard"
    summary["A1"].font = Font(name="Calibri", size=16, bold=True, color="1B2B4B")
    summary["A3"] = "Tracker updated"
    summary["B3"] = datetime.now().strftime("%d-%b-%Y %H:%M")

    metrics = [
        ("Total Enquiries", '=COUNTA(\'Enquiry Tracker\'!D2:D201)+COUNTA(\'Apply Now\'!D2:D201)'),
        ("New Leads", '=COUNTIF(\'Enquiry Tracker\'!L2:L201,"New")+COUNTIF(\'Apply Now\'!M2:M201,"New")'),
        ("Follow-up Due", '=COUNTIF(\'Enquiry Tracker\'!L2:L201,"Follow-up Due")+COUNTIF(\'Apply Now\'!M2:M201,"Follow-up Due")'),
        ("Contacted", '=COUNTIF(\'Enquiry Tracker\'!O2:O201,"Yes")+COUNTIF(\'Apply Now\'!P2:P201,"Yes")'),
        ("Interested", '=COUNTIF(\'Enquiry Tracker\'!L2:L201,"Interested")+COUNTIF(\'Apply Now\'!M2:M201,"Interested")'),
        ("Enrolled", '=COUNTIF(\'Enquiry Tracker\'!Q2:Q201,"Enrolled")+COUNTIF(\'Apply Now\'!R2:R201,"Enrolled")'),
    ]

    for idx, (label, formula) in enumerate(metrics, start=5):
        summary[f"A{idx}"] = label
        summary[f"A{idx}"].font = Font(bold=True, color="FFFFFF")
        summary[f"A{idx}"].fill = PatternFill("solid", fgColor="C6923D")
        summary[f"B{idx}"] = formula
        summary[f"B{idx}"].fill = PatternFill("solid", fgColor="F8F1E3")
        summary[f"B{idx}"].font = Font(bold=True, color="1B2B4B")
        style_body(summary[f"A{idx}"])
        style_body(summary[f"B{idx}"])

    summary["D5"] = "Recommended workflow"
    summary["D5"].font = Font(bold=True, color="1B2B4B")
    workflow = [
        "1. Check new enquiries each morning and evening.",
        "2. Mark Contacted? as Yes once someone from the team calls or replies.",
        "3. Fill Outcome and Next Action after every conversation.",
        "4. Use Follow-up Date so no lead is missed.",
        "5. Move final cases to Enrolled or Closed.",
    ]
    for offset, line in enumerate(workflow, start=6):
        summary[f"D{offset}"] = line

    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 16
    summary.column_dimensions["D"].width = 48

    output_path = Path(__file__).resolve().parent / OUTPUT_FILE
    workbook.save(output_path)
    print(output_path)


if __name__ == "__main__":
    main()
